"""Excel (.xlsx) workbook output for the cost estimator.

Five sheets: token estimate, recommendation, ranked pricing comparison,
all kept providers, and excluded records.
"""
from __future__ import annotations

from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .cost_engine import CostLine
from .recommender import ScoredLine
from .token_estimator import TokenEstimate

HEADER_FILL = PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14)
HEADER_FONT = Font(bold=True)
MONEY_FORMAT = '"$"#,##0.000000'
SCORE_FORMAT = "0.0000"
PCT_FORMAT = "0.00%"
INT_FORMAT = "#,##0"


def _autosize(ws: Worksheet, widths: Sequence[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_header(ws: Worksheet, row: int, values: Sequence[str]) -> None:
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _money(cell, value: float | int | None) -> None:
    if value is None:
        cell.value = None
        return
    cell.value = float(value)
    cell.number_format = MONEY_FORMAT


def _int(cell, value: int | None) -> None:
    cell.value = value
    cell.number_format = INT_FORMAT


def _score(cell, value: float | None) -> None:
    if value is None:
        cell.value = None
        return
    cell.value = float(value)
    cell.number_format = SCORE_FORMAT


def _title(ws: Worksheet, text: str) -> None:
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = TITLE_FONT


def _add_token_estimate_sheet(
    wb: Workbook, prompt: str, est: TokenEstimate, pricing_meta: dict[str, Any]
) -> None:
    ws = wb.create_sheet("Token Estimate")
    _title(ws, "AI Token & Price Estimator - Token Estimate")

    ws.cell(row=3, column=1, value="Generated").font = HEADER_FONT
    ws.cell(row=3, column=2, value=pricing_meta.get("fetched_at", ""))
    ws.cell(row=4, column=1, value="Pricing source").font = HEADER_FONT
    ws.cell(row=4, column=2, value=str(pricing_meta.get("source", "")))
    ws.cell(row=5, column=1, value="Stale").font = HEADER_FONT
    ws.cell(row=5, column=2, value=bool(pricing_meta.get("stale", False)))

    ws.cell(row=7, column=1, value="Project prompt").font = HEADER_FONT
    prompt_cell = ws.cell(row=8, column=1, value=prompt)
    prompt_cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.cell(row=10, column=1, value="Token counts").font = HEADER_FONT
    _write_header(ws, 11, ["Metric", "Tokens"])
    rows = [
        ("Task kind", est.task_kind),
        ("Input tokens", est.input_tokens),
        ("Output tokens (estimated)", est.output_tokens_estimated),
        ("Reasoning buffer", est.reasoning_buffer_tokens),
        ("System overhead", est.system_overhead_tokens),
        ("Total", est.total_tokens),
        ("Multiplier used", est.multiplier_used),
    ]
    r = 12
    for label, value in rows:
        ws.cell(row=r, column=1, value=label)
        c = ws.cell(row=r, column=2)
        if isinstance(value, (int,)) and not isinstance(value, bool):
            _int(c, value)
        else:
            c.value = value
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Notes").font = HEADER_FONT
    r += 1
    for note in est.notes:
        ws.cell(row=r, column=1, value=f"- {note}")
        r += 1

    _autosize(ws, [32, 24, 80])
    ws.row_dimensions[8].height = 60


def _add_recommendation_sheet(wb: Workbook, recs: list[ScoredLine]) -> None:
    ws = wb.create_sheet("Recommendation")
    _title(ws, "Recommendation (weighted scoring)")

    ws.cell(row=3, column=1, value="Weights: cost=0.40 capability=0.30 context=0.20 latency=0.10").font = HEADER_FONT

    if not recs:
        ws.cell(row=5, column=1, value="No recommendations available.")
        _autosize(ws, [80])
        return

    best = recs[0]
    ws.cell(row=5, column=1, value="★ BEST").font = TITLE_FONT
    rows = [
        ("Provider", best.cost_line.provider),
        ("Model", best.cost_line.model),
        ("Total cost (USD)", best.cost_line.total_cost),
        ("Weighted score", best.weighted_score),
        ("Cost score", best.cost_score),
        ("Capability score", best.capability_score),
        ("Context score", best.context_score),
        ("Latency score", best.latency_score),
        ("Fits context window", "yes" if best.cost_line.fits_context_window else "NO"),
        ("Context used", f"{best.cost_line.input_tokens}/{best.cost_line.context_window}"),
        ("Reason", best.reason),
    ]
    r = 6
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = HEADER_FONT
        c = ws.cell(row=r, column=2)
        if label == "Total cost (USD)":
            _money(c, value)
        elif label == "Weighted score" or label.endswith("score"):
            _score(c, value)
        else:
            c.value = value
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Alternatives").font = HEADER_FONT
    r += 1
    _write_header(ws, r, ["Rank", "Provider", "Model", "Total cost (USD)", "Weighted score", "Cost score", "Reason"])
    r += 1
    for s in recs[1:]:
        ws.cell(row=r, column=1, value=s.rank)
        ws.cell(row=r, column=2, value=s.cost_line.provider)
        ws.cell(row=r, column=3, value=s.cost_line.model)
        _money(ws.cell(row=r, column=4), s.cost_line.total_cost)
        _score(ws.cell(row=r, column=5), s.weighted_score)
        _score(ws.cell(row=r, column=6), s.cost_score)
        ws.cell(row=r, column=7, value=s.reason)
        r += 1

    _autosize(ws, [22, 14, 40, 16, 14, 12, 60])


def _add_pricing_comparison_sheet(wb: Workbook, ranked: list[CostLine]) -> None:
    ws = wb.create_sheet("Pricing Comparison")
    _title(ws, "Pricing Comparison (ranked by total cost)")

    headers = [
        "Rank", "Provider", "Model", "Context window",
        "Input cost / 1M", "Output cost / 1M",
        "Input cost", "Output cost", "Total cost",
        "Fits context", "Latency p50 (ms)", "Capability tags",
    ]
    _write_header(ws, 3, headers)

    r = 4
    for i, ln in enumerate(ranked, 1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=ln.provider)
        ws.cell(row=r, column=3, value=ln.model)
        _int(ws.cell(row=r, column=4), ln.context_window)
        _money(ws.cell(row=r, column=5), ln.input_cost_per_1m)
        _money(ws.cell(row=r, column=6), ln.output_cost_per_1m)
        _money(ws.cell(row=r, column=7), ln.input_cost)
        _money(ws.cell(row=r, column=8), ln.output_cost)
        _money(ws.cell(row=r, column=9), ln.total_cost)
        ws.cell(row=r, column=10, value="yes" if ln.fits_context_window else "NO")
        _int(ws.cell(row=r, column=11), ln.latency_p50_ms)
        ws.cell(row=r, column=12, value=", ".join(ln.capability_tags))
        r += 1

    _autosize(ws, [6, 14, 40, 14, 14, 14, 14, 14, 14, 12, 14, 32])


def _add_providers_sheet(wb: Workbook, kept_records: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Providers")
    _title(ws, "All kept providers (post-filter)")

    headers = [
        "Provider", "Model", "Context window",
        "Input cost / 1M", "Output cost / 1M",
        "Reasoning", "Coding", "Creativity",
        "Latency p50 (ms)", "Capability tags", "Source URL", "Fetched at",
    ]
    _write_header(ws, 3, headers)

    r = 4
    for rec in kept_records:
        ws.cell(row=r, column=1, value=rec.get("provider", ""))
        ws.cell(row=r, column=2, value=rec.get("model", ""))
        _int(ws.cell(row=r, column=3), rec.get("context_window"))
        _money(ws.cell(row=r, column=4), rec.get("input_cost_per_1m"))
        _money(ws.cell(row=r, column=5), rec.get("output_cost_per_1m"))
        _score(ws.cell(row=r, column=6), rec.get("reasoning_strength"))
        _score(ws.cell(row=r, column=7), rec.get("coding_strength"))
        _score(ws.cell(row=r, column=8), rec.get("creativity_strength"))
        _int(ws.cell(row=r, column=9), rec.get("latency_p50_ms"))
        ws.cell(row=r, column=10, value=", ".join(rec.get("capability_tags", []) or []))
        ws.cell(row=r, column=11, value=rec.get("source_url", ""))
        ws.cell(row=r, column=12, value=rec.get("fetched_at", ""))
        r += 1

    _autosize(ws, [14, 40, 14, 14, 14, 12, 12, 12, 14, 32, 50, 22])


def _add_excluded_sheet(wb: Workbook, excluded: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Excluded")
    _title(ws, "Excluded records (filtered out)")

    if not excluded:
        ws.cell(row=3, column=1, value="No records were excluded.")
        _autosize(ws, [40])
        return

    _write_header(ws, 3, ["Provider", "Model", "Reasons"])
    r = 4
    for e in excluded:
        ws.cell(row=r, column=1, value=e.get("provider", ""))
        ws.cell(row=r, column=2, value=e.get("model", ""))
        ws.cell(row=r, column=3, value=", ".join(e.get("reasons", []) or []))
        r += 1

    _autosize(ws, [14, 40, 80])


def build_workbook(
    prompt: str,
    est: TokenEstimate,
    pricing_meta: dict[str, Any],
    kept_records: list[dict[str, Any]],
    excluded: list[dict[str, Any]],
    ranked: list[CostLine],
    recs: list[ScoredLine],
) -> Workbook:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    _add_token_estimate_sheet(wb, prompt, est, pricing_meta)
    _add_recommendation_sheet(wb, recs)
    _add_pricing_comparison_sheet(wb, ranked)
    _add_providers_sheet(wb, kept_records)
    _add_excluded_sheet(wb, excluded)

    return wb
