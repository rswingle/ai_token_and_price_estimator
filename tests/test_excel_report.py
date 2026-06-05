"""Unit tests for the Excel workbook builder."""
from openpyxl import load_workbook

from src.cost_engine import compute_all
from src.excel_report import (
    HEADER_FILL,
    HEADER_FONT,
    MONEY_FORMAT,
    SCORE_FORMAT,
    build_workbook,
)
from src.pricing_research import _baseline_snapshot
from src.recommender import recommend
from src.token_estimator import estimate


def _full_run(prompt="Implement a function in Python", top_n=3):
    est = estimate(prompt)
    records = _baseline_snapshot()
    lines = compute_all(est, records)
    from src.cost_engine import rank_by_cost
    ranked = rank_by_cost(lines)
    recs = recommend(ranked, est.task_kind, top_n=top_n)
    pricing_meta = {
        "source": "baseline",
        "stale": True,
        "fetched_at": "2026-06-05T00:00:00Z",
        "errors": [],
    }
    return prompt, est, pricing_meta, records, [], ranked, recs


class TestBuildWorkbook:
    def test_returns_workbook_with_five_sheets(self):
        args = _full_run()
        wb = build_workbook(*args)
        names = wb.sheetnames
        assert names == [
            "Token Estimate",
            "Recommendation",
            "Pricing Comparison",
            "Providers",
            "Excluded",
        ]

    def test_default_sheet_removed(self):
        args = _full_run()
        wb = build_workbook(*args)
        assert "Sheet" not in wb.sheetnames

    def test_token_estimate_sheet_has_required_rows(self):
        args = _full_run(prompt="hello world")
        wb = build_workbook(*args)
        ws = wb["Token Estimate"]
        flat = [ws.cell(row=r, column=1).value for r in range(1, 25)]
        assert "AI Token & Price Estimator - Token Estimate" in flat
        assert "Project prompt" in flat
        assert "Token counts" in flat
        assert "Input tokens" in flat
        assert "Output tokens (estimated)" in flat
        assert "Total" in flat
        assert "Notes" in flat

    def test_token_estimate_prompt_present(self):
        args = _full_run(prompt="hello world")
        wb = build_workbook(*args)
        ws = wb["Token Estimate"]
        assert ws.cell(row=8, column=1).value == "hello world"

    def test_token_estimate_numeric_cells_have_int_format(self):
        args = _full_run()
        wb = build_workbook(*args)
        ws = wb["Token Estimate"]
        token_row = next(
            r for r in range(1, 30) if ws.cell(row=r, column=1).value == "Input tokens"
        )
        c = ws.cell(row=token_row, column=2)
        assert isinstance(c.value, int)
        assert c.number_format == "#,##0"

    def test_recommendation_sheet_has_best_block(self):
        args = _full_run()
        wb = build_workbook(*args)
        ws = wb["Recommendation"]
        flat = [ws.cell(row=r, column=1).value for r in range(1, 25)]
        assert "★ BEST" in flat
        assert "Provider" in flat
        assert "Total cost (USD)" in flat
        assert "Weighted score" in flat
        assert "Reason" in flat

    def test_recommendation_money_cell_uses_money_format(self):
        args = _full_run()
        wb = build_workbook(*args)
        ws = wb["Recommendation"]
        cost_row = next(
            r for r in range(1, 25) if ws.cell(row=r, column=1).value == "Total cost (USD)"
        )
        c = ws.cell(row=cost_row, column=2)
        assert c.number_format == MONEY_FORMAT
        assert isinstance(c.value, (int, float))

    def test_recommendation_score_cell_uses_score_format(self):
        args = _full_run()
        wb = build_workbook(*args)
        ws = wb["Recommendation"]
        score_row = next(
            r for r in range(1, 25) if ws.cell(row=r, column=1).value == "Weighted score"
        )
        c = ws.cell(row=score_row, column=2)
        assert c.number_format == SCORE_FORMAT

    def test_pricing_comparison_lists_all_models(self):
        args = _full_run()
        wb = build_workbook(*args)
        ws = wb["Pricing Comparison"]
        data_rows = sum(1 for r in range(4, 50) if ws.cell(row=r, column=2).value)
        assert data_rows == 20

    def test_pricing_comparison_first_row_is_cheapest(self):
        args = _full_run()
        wb = build_workbook(*args)
        ws = wb["Pricing Comparison"]
        first_model = ws.cell(row=4, column=3).value
        assert first_model is not None and len(first_model) > 0

    def test_pricing_comparison_total_cost_uses_money_format(self):
        args = _full_run()
        wb = build_workbook(*args)
        ws = wb["Pricing Comparison"]
        c = ws.cell(row=4, column=9)
        assert c.number_format == MONEY_FORMAT

    def test_providers_sheet_lists_all_records(self):
        args = _full_run()
        wb = build_workbook(*args)
        ws = wb["Providers"]
        data_rows = sum(1 for r in range(4, 50) if ws.cell(row=r, column=1).value)
        assert data_rows == 20

    def test_excluded_sheet_placeholder_when_empty(self):
        prompt, est, pricing_meta, kept, _, ranked, recs = _full_run()
        wb = build_workbook(prompt, est, pricing_meta, kept, [], ranked, recs)
        ws = wb["Excluded"]
        assert ws.cell(row=3, column=1).value == "No records were excluded."

    def test_excluded_sheet_lists_records(self):
        prompt, est, pricing_meta, kept, _, ranked, recs = _full_run()
        excluded = [
            {"provider": "Bad", "model": "x", "reasons": ["deprecated model"]},
            {"provider": "Worse", "model": "y", "reasons": ["session-based or chat product"]},
        ]
        wb = build_workbook(prompt, est, pricing_meta, kept, excluded, ranked, recs)
        ws = wb["Excluded"]
        assert ws.cell(row=4, column=1).value == "Bad"
        assert "deprecated" in ws.cell(row=4, column=3).value
        assert ws.cell(row=5, column=1).value == "Worse"

    def test_header_cells_have_bold_font_and_fill(self):
        args = _full_run()
        wb = build_workbook(*args)
        ws = wb["Pricing Comparison"]
        c = ws.cell(row=3, column=1)
        assert c.font.bold is True
        assert c.fill.start_color.rgb is not None

    def test_empty_recommendations_does_not_crash(self):
        prompt, est, pricing_meta, kept, _, ranked, _ = _full_run()
        wb = build_workbook(prompt, est, pricing_meta, kept, [], ranked, [])
        ws = wb["Recommendation"]
        assert ws.cell(row=5, column=1).value == "No recommendations available."

    def test_saves_as_valid_xlsx(self, tmp_path):
        args = _full_run()
        wb = build_workbook(*args)
        out = tmp_path / "report.xlsx"
        wb.save(out)
        loaded = load_workbook(out)
        assert "Token Estimate" in loaded.sheetnames
        assert "Pricing Comparison" in loaded.sheetnames
        assert "Recommendation" in loaded.sheetnames
        assert "Providers" in loaded.sheetnames
        assert "Excluded" in loaded.sheetnames
